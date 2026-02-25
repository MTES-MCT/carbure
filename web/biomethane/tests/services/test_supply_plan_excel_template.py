"""Tests for the supply plan Excel template creation service."""

from django.test import TestCase
from openpyxl import load_workbook

from biomethane.services.supply_plan_excel_template import create_supply_plan_template


class SupplyPlanExcelTemplateTests(TestCase):
    """Tests for supply plan Excel template generation."""

    fixtures = ["json/departments.json", "json/countries.json"]

    def setUp(self):
        self.file = create_supply_plan_template()
        self.workbook = load_workbook(self.file)

    def test_create_supply_plan_template_creates_file(self):
        """Test that the Excel template is created successfully."""
        file = create_supply_plan_template()

        self.assertIsNotNone(file)
        self.assertTrue(file.readable())

    def test_template_has_all_required_sheets(self):
        """Test that all required sheets are present."""
        expected_sheets = [
            "Plan d'approvisionnement",
            "Departements",
            "Pays",
        ]

        actual_sheets = self.workbook.sheetnames

        for sheet_name in expected_sheets:
            with self.subTest(sheet_name=sheet_name):
                self.assertIn(sheet_name, actual_sheets, f"La feuille '{sheet_name}' devrait être présente")

    def test_main_sheet_has_correct_columns(self):
        """Test that the main sheet has the correct columns."""
        main_sheet = self.workbook["Plan d'approvisionnement"]

        expected_columns = [
            "Intrant",
            "Type de CIVE",
            "Précisez la culture",
            "Type de collecte",
            "Unité",
            "Ratio de matière sèche (%)",
            "Volume (tMB ou tMS)",
            "Département",
            "Distance moyenne pondérée (km)",
            "Distance maximale (km)",
            "Pays d'origine",
        ]

        # Read the first row (headers) and filter out None values
        actual_columns = [cell.value for cell in main_sheet[1] if cell.value is not None]

        self.assertEqual(actual_columns, expected_columns)

    def test_departments_sheet_has_data(self):
        """Test that the Departments sheet contains data."""
        departments_sheet = self.workbook["Departements"]

        # Check that there is data
        self.assertGreater(departments_sheet.max_row, 1, "La feuille Departements devrait contenir des données")

        # Check some key departments
        codes = [departments_sheet.cell(i, 1).value for i in range(2, departments_sheet.max_row + 1)]
        self.assertIn("75", codes)  # Paris
        self.assertIn("2A", codes)  # Corse-du-Sud
        self.assertIn("971", codes)  # Guadeloupe

    def test_countries_sheet_has_data(self):
        """Test that the Countries sheet contains data."""
        countries_sheet = self.workbook["Pays"]

        # Check that there is data (at least France)
        self.assertGreater(countries_sheet.max_row, 1)

        # Check headers
        self.assertEqual(countries_sheet.cell(1, 1).value, "Code")
        self.assertEqual(countries_sheet.cell(1, 2).value, "Nom")
