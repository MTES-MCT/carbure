import os

from django.test import SimpleTestCase
from openpyxl import load_workbook

from core.import_export_template import create_import_template, to_snake_case


class ImportExportTemplateTests(SimpleTestCase):
    def _load_workbook(self, title, columns):
        file_handle = create_import_template(title=title, columns=columns)
        self.addCleanup(self._cleanup_template, file_handle.name)
        file_handle.close()
        return load_workbook(file_handle.name)

    @staticmethod
    def _cleanup_template(path):
        if os.path.exists(path):
            os.remove(path)

    def test_create_import_template_builds_sheets_headers_and_validations(self):
        columns = [
            {"header": "Name"},
            {"header": "Status", "options": ["Active", "Inactive"]},
            {"header": "Type", "options": ["A", "B", "C"]},
        ]

        workbook = self._load_workbook("My Template", columns)
        self.addCleanup(workbook.close)

        main_sheet = workbook["My Template"]
        reference_sheet = workbook["References"]

        self.assertEqual([cell.value for cell in main_sheet[1]], ["Name", "Status", "Type"])
        self.assertEqual(reference_sheet.sheet_state, "hidden")
        self.assertTrue(reference_sheet.protection.sheet)
        self.assertEqual([cell.value for cell in reference_sheet[1]], ["Name", "Status", "Type"])
        self.assertEqual(reference_sheet["B2"].value, "Active")
        self.assertEqual(reference_sheet["B3"].value, "Inactive")
        self.assertEqual(reference_sheet["C2"].value, "A")
        self.assertEqual(reference_sheet["C4"].value, "C")

        validations = main_sheet.data_validations.dataValidation
        self.assertEqual(len(validations), 2)

        def normalize_formula(formula):
            if formula and not formula.startswith("="):
                return f"={formula}"
            return formula

        validation_data = {(normalize_formula(dv.formula1), str(dv.sqref)) for dv in validations}
        self.assertEqual(
            validation_data,
            {
                ("=References!$B$2:$B$3", "B2:B1000"),
                ("=References!$C$2:$C$4", "C2:C1000"),
            },
        )

    def test_create_import_template_skips_validation_when_no_options(self):
        columns = [{"header": "Name"}]
        workbook = self._load_workbook("Only Name", columns)
        self.addCleanup(workbook.close)

        main_sheet = workbook["Only Name"]
        self.assertEqual(len(main_sheet.data_validations.dataValidation), 0)

    def test_create_import_template_uses_snake_case_filename(self):
        file_handle = create_import_template("My Fancy-Title!", columns=[])
        self.addCleanup(self._cleanup_template, file_handle.name)
        file_handle.close()

        self.assertEqual(os.path.basename(file_handle.name), "my_fancy_title_import_template.xlsx")

    def test_to_snake_case_handles_punctuation_and_camel_case(self):
        self.assertEqual(to_snake_case("My Fancy-Title!"), "my_fancy_title")
        self.assertEqual(to_snake_case("HelloWorld"), "hello_world")
