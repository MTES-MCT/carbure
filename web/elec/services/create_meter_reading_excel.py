from datetime import date, timedelta
from typing import Iterable, Optional, TypedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from elec.models import ElecChargePoint, ElecMeterReadingApplication
from elec.repositories.meter_reading_repository import MeterReadingRepository


class MeterReadingData(TypedDict):
    charge_point_id: str
    previous_reading: float
    current_reading: Optional[float]
    reading_date: date


def create_meter_readings_excel(
    name: str,
    quarter: int,
    year: int,
    meter_readings_data: list[MeterReadingData],
    extended=False,
):
    workbook = Workbook()
    sheet = workbook.active

    last_day = (date(year, quarter * 3, 1) + timedelta(days=31)).replace(day=1)
    renewable_share = MeterReadingRepository.get_renewable_share(year)

    # HEADER

    sheet["A1"] = "Identifiant du point de recharge communiqué à transport.data.gouv"
    sheet["B1"] = "Energie active totale soutirée lors du relevé précédent"
    sheet["C1"] = "Energie active totale soutirée au " + last_day.strftime("%d/%m/%Y")
    sheet["D1"] = "Date du relevé (JJ/MM/AAAA)"

    if extended:
        sheet["E1"] = "Electricité consommée sur la période"
        sheet["F1"] = "Electricité renouvelable consommée sur la période"

        sheet["H1"] = "Part renouvelable de l'électricité sur la période"
        sheet["H2"] = f"{renewable_share * 100}%"

    # BODY

    stations = set()

    for i, reading in enumerate(meter_readings_data, 2):
        stations.add(reading.get("charge_point_id")[0:5])
        reading_date = reading.get("reading_date")
        sheet[f"A{i}"] = reading["charge_point_id"]
        sheet[f"B{i}"] = reading["previous_reading"]
        sheet[f"C{i}"] = reading.get("current_reading") or ""
        sheet[f"D{i}"] = reading_date.strftime("%d/%m/%Y") if reading_date else ""

        if extended:
            sheet[f"E{i}"] = f"=IF(ISBLANK(C{i}), 0, C{i} - B{i})"
            sheet[f"F{i}"] = f"=IF(ISBLANK(E{i}), 0, ROUND(E{i} * {renewable_share}, 3))"

    # FOOTER

    last_reading_row = 1 + len(meter_readings_data)
    total_row = last_reading_row + 2
    sheet[f"A{total_row}"] = "Total"
    sheet[f"C{total_row}"] = f"=SUM(C2:C{total_row - 1})"

    last_row = total_row + 2 + len(stations)

    if extended:
        sheet[f"E{total_row}"] = f"=SUM(E2:E{total_row - 1})"
        sheet[f"F{total_row}"] = f"=SUM(F2:F{total_row - 1})"

        sheet[f"A{total_row + 2}"] = "Unités d'exploitation:"
        for i, station in enumerate(stations, 1):
            sheet[f"A{total_row + 2 + i}"] = station

    # STYLING

    bold = Font(bold=True)
    side = Side(border_style="thin", color="AAAAAA")
    border = Border(top=side, left=side, right=side, bottom=side)
    yellow_bg = PatternFill(start_color="FFF2CC", fill_type="solid")
    blue_bg = PatternFill(start_color="D9E1F2", fill_type="solid")

    # header text
    for cell in sheet[1]:
        cell.font = bold
        cell.alignment = Alignment(wrapText=True)

    # row styles
    for i in range(1, last_row):
        sheet.row_dimensions[i].height = 48 if i == 1 else 16

        if i <= last_reading_row:
            sheet[f"A{i}"].fill = yellow_bg
            sheet[f"A{i}"].border = border
            sheet[f"B{i}"].fill = blue_bg
            sheet[f"B{i}"].border = border
            sheet[f"C{i}"].fill = blue_bg
            sheet[f"C{i}"].border = border
            sheet[f"D{i}"].fill = blue_bg
            sheet[f"D{i}"].border = border

    # default column width
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 24

    # total row in bold
    sheet[total_row][0].font = bold
    sheet[total_row][2].font = bold
    sheet[total_row][3].font = bold

    if extended:
        sheet[total_row][4].font = bold

        # station header in bold
        sheet[f"A{total_row + 2}"].font = bold

    file_path = f"/tmp/{name}.xlsx"
    workbook.save(file_path)
    return open(file_path, "rb")


def create_meter_readings_data(
    charge_points: Iterable[ElecChargePoint],
    previous_application: ElecMeterReadingApplication,
    current_readings: list[dict] = [],
):
    previous_readings_by_charge_point = get_previous_readings_by_charge_point(charge_points, previous_application)
    current_readings_by_charge_point = get_current_readings_by_charge_point(current_readings)

    meter_reading_data: list[MeterReadingData] = []
    for charge_point in charge_points:
        if charge_point.is_article_2:
            continue

        charge_point_id = charge_point.charge_point_id
        current_reading = current_readings_by_charge_point.get(charge_point_id)

        if len(current_readings) > 0 and current_reading is None:
            continue

        reading_data = {
            "charge_point_id": charge_point_id,
            "previous_reading": previous_readings_by_charge_point.get(charge_point_id),
            "current_reading": current_reading.get("extracted_energy") if current_reading else 0,
            "reading_date": current_reading.get("reading_date") if current_reading else 0,
        }

        meter_reading_data.append(reading_data)

    return meter_reading_data


def get_previous_readings_by_charge_point(
    charge_points: Iterable[ElecChargePoint],
    previous_application: ElecMeterReadingApplication,
):
    previous_readings_by_charge_point = {}

    # initialize previous readings using the first one set during the charge point registration
    for charge_point in charge_points:
        previous_readings_by_charge_point[charge_point.charge_point_id] = charge_point.measure_energy

    # then if there was a previous registration, use its data to specify the previous reading latest value
    if previous_application:
        for reading in previous_application.elec_meter_readings.all():
            previous_readings_by_charge_point[reading.charge_point.charge_point_id] = reading.extracted_energy

    return previous_readings_by_charge_point


def get_current_readings_by_charge_point(current_readings: list[dict]):
    current_readings_by_charge_point: dict[str, dict] = {}
    for reading in current_readings:
        current_readings_by_charge_point[reading.get("charge_point_id")] = reading

    return current_readings_by_charge_point
