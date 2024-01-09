import traceback
from datetime import date
from typing import Iterable
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from django.http import HttpResponse
from django.views.decorators.http import require_GET
from core.common import ErrorResponse
from core.decorators import check_user_rights
from core.models import UserRights
from elec.models.elec_charge_point import ElecChargePoint
from elec.models.elec_charge_point_application import ElecChargePointApplication
from elec.models.elec_meter_reading_application import ElecMeterReadingApplication


class ApplicationTemplateError:
    TOO_LATE = "TOO_LATE"
    NO_CHARGE_POINT_AVAILABLE = "NO_CHARGE_POINT_AVAILABLE"
    TEMPLATE_CREATION_FAILED = "TEMPLATE_CREATION_FAILED"


@require_GET
@check_user_rights(role=[UserRights.ADMIN, UserRights.RW])
def get_application_template(request, entity):
    today = date.today()
    first_day = first_day_of_current_quarter()

    if (today - first_day).days > 20:
        return ErrorResponse(400, ApplicationTemplateError.TOO_LATE)

    try:
        charge_points = (
            ElecChargePoint.objects.select_related("application")
            .filter(cpo=entity, application__status=ElecChargePointApplication.ACCEPTED)
            .order_by("charge_point_id")
        )

        if charge_points.count() == 0:
            return ErrorResponse(400, ApplicationTemplateError.NO_CHARGE_POINT_AVAILABLE)

        last_application = (
            ElecMeterReadingApplication.objects.filter(cpo=entity)
            .order_by("-year", "-quarter")
            .prefetch_related("elec_meter_readings")
            .first()
        )

        workbook = create_meter_reading_template(
            first_day,
            charge_points,
            last_application,
        )
    except:
        traceback.print_exc()
        return ErrorResponse(ApplicationTemplateError.TEMPLATE_CREATION_FAILED)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=template.xlsx"
    workbook.save(response)

    return response


def create_meter_reading_template(
    last_day: date,
    charge_points: Iterable[ElecChargePoint],
    last_application: ElecMeterReadingApplication,
):
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Identifiant du point de recharge communiqué à transport.data.gouv"
    sheet["B1"] = "Energie active totale soutirée lors du relevé précédent"
    sheet["C1"] = "Energie active totale soutirée au " + last_day.strftime("%d/%m/%Y")
    sheet["D1"] = "Electricité consommée sur la période"
    sheet["E1"] = "Electricité renouvelable consommée sur la période"

    sheet["G1"] = "Part renouvelable de l'électricité sur la période"
    sheet["G2"] = "24,92%"

    last_readings_by_charge_point = {}

    if last_application:
        for reading in last_application.elec_meter_readings:
            last_readings_by_charge_point[reading.charge_point.charge_point_id] = reading.extracted_energy
    else:
        for charge_point in charge_points:
            last_readings_by_charge_point[charge_point.charge_point_id] = charge_point.measure_energy

    stations = set()
    renewable_part = 0.2492

    for i, charge_point in enumerate(charge_points, 2):
        stations.add(charge_point.charge_point_id[0:5])
        sheet[f"A{i}"] = charge_point.charge_point_id
        sheet[f"B{i}"] = last_readings_by_charge_point.get(charge_point.charge_point_id, 0)
        sheet[f"C{i}"] = ""
        sheet[f"D{i}"] = f"=IF(ISBLANK(C{i}), 0, C{i} - B{i})"
        sheet[f"E{i}"] = f"=IF(ISBLANK(D{i}), 0, ROUND(D{i} * {renewable_part}, 2))"

    last_charge_point_row = 1 + len(charge_points)
    total_row = last_charge_point_row + 2
    sheet[f"A{total_row}"] = "Total"
    sheet[f"D{total_row}"] = f"=SUM(D2:D{total_row - 1})"
    sheet[f"E{total_row}"] = f"=SUM(E2:E{total_row - 1})"

    sheet[f"A{total_row + 2}"] = "Stations:"
    last_row = total_row + 2 + len(stations)
    for i, station in enumerate(stations, 1):
        sheet[f"A{total_row + 2 + i}"] = station

    # STYLING

    bold = Font(bold=True)
    side = Side(border_style="thin", color="AAAAAA")
    border = Border(top=side, left=side, right=side, bottom=side)
    yellow_bg = PatternFill(start_color="FFF2CC", fill_type="solid")
    blue_bg = PatternFill(start_color="D9E1F2", fill_type="solid")

    # header text
    for cell in sheet[1]:
        cell.font = bold
        cell.alignment = Alignment(wrapText=True)

    # row styles
    for i in range(1, last_row):
        sheet.row_dimensions[i].height = 48 if i == 1 else 16

        if i <= last_charge_point_row:
            sheet[f"A{i}"].fill = yellow_bg
            sheet[f"A{i}"].border = border
            sheet[f"B{i}"].fill = blue_bg
            sheet[f"B{i}"].border = border
            sheet[f"C{i}"].fill = blue_bg
            sheet[f"C{i}"].border = border

    # default column width
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 24

    # total row in bold
    sheet[total_row][0].font = bold
    sheet[total_row][3].font = bold
    sheet[total_row][4].font = bold

    # station header in bold
    sheet[f"A{total_row + 2}"].font = bold

    return workbook


def first_day_of_current_quarter():
    now = date.today()
    quarter = (now.month - 1) // 3 + 1
    month = (quarter - 1) * 3 + 1
    return date(now.year, month, 1)
