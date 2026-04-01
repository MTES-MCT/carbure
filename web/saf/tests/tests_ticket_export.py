from io import BytesIO

from django.urls import reverse
from openpyxl import load_workbook

from saf.models import SafTicket
from saf.tests import TestCase


class SafTicketExportTest(TestCase):
    def test_saf_ticket_export_returns_valid_excel_file(self):
        self.ticket.client = self.airline
        self.ticket.ets_status = SafTicket.ETS_VALUATION
        self.ticket.save()

        response = self.client.get(reverse("saf-tickets-export"), {"entity_id": self.airline.id})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.ms-excel")
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        self.addCleanup(workbook.close)

        self.assertEqual(workbook.sheetnames, ["tickets", "aeroports", "biocarburants"])

        tickets_sheet = workbook["tickets"]
        headers = [cell.value for cell in tickets_sheet[1]]

        self.assertEqual(
            headers,
            [
                "carbure_id",
                "year",
                "assignment_period",
                "agreement_reference",
                "agreement_date",
                "volume",
                "biofuel",
                "feedstock",
                "country_of_origin",
                "supplier",
                "client",
                "client_type",
                "producer",
                "production_site",
                "production_country",
                "production_site_commissioning_date",
                "origin_depot",
                "reception_airport",
                "eec",
                "el",
                "ep",
                "etd",
                "eu",
                "esca",
                "eccs",
                "eccr",
                "eee",
                "ghg_total",
                "ghg_reduction",
                "free_field",
                "ets_status",
            ],
        )

        self.assertEqual(tickets_sheet["A2"].value, self.ticket.carbure_id)
        self.assertEqual(tickets_sheet["B2"].value, self.ticket.year)
        self.assertEqual(tickets_sheet["F2"].value, self.ticket.volume)
        self.assertEqual(tickets_sheet["AE2"].value, SafTicket.ETS_VALUATION)

    def test_saf_ticket_export_does_not_include_ets_status_for_operator(self):
        self.ticket.ets_status = SafTicket.ETS_VALUATION
        self.ticket.save()

        response = self.client.get(reverse("saf-tickets-export"), {"entity_id": self.entity.id})

        workbook = load_workbook(BytesIO(response.content))
        self.addCleanup(workbook.close)

        tickets_sheet = workbook["tickets"]
        headers = [cell.value for cell in tickets_sheet[1]]

        self.assertNotIn("ets_status", headers)
