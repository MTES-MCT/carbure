from calendar import c
import re
import traceback
from typing import List, TypedDict, Tuple
from openpyxl import Workbook, load_workbook
from core.common import CarbureException

from doublecount.dc_sanity_checks import DoubleCountingError


class SourcingRow(TypedDict):
    line: int
    year: int
    feedstock: str | None
    origin_country: str | None
    supply_country: str | None
    transit_country: str | None
    metric_tonnes: int


class ProductionRow(TypedDict):
    line: int
    year: int
    feedstock: str | None
    feedstock_check: str | None
    biofuel: str | None
    max_production_capacity: int
    estimated_production: int
    requested_quota: int


class ProductionBaseRow(TypedDict):
    line: int
    year: int
    feedstock: str | None
    biofuel: str | None


class RequestedQuotaRow(ProductionBaseRow):
    requested_quota: int


class ProductionForecastRow(ProductionBaseRow):
    estimated_production: int


class ProductionMaxRow(ProductionBaseRow):
    max_production_capacity: int


def parse_dc_excel(
    filename: str,
) -> Tuple[dict[str], List[SourcingRow], List[ProductionMaxRow], List[ProductionForecastRow], List[RequestedQuotaRow]]:
    excel_file = load_workbook(filename, data_only=True)

    info = parse_info(excel_file)
    requested_quota_rows = parse_requested_quota(excel_file)

    years = [production_row["year"] for production_row in requested_quota_rows]
    start_year = max(years) - 1 if len(years) > 0 else info["start_year"] + 1 if info["start_year"] else 0

    sourcing_forecast_rows = parse_sourcing_forecast(excel_file, requested_year=start_year)
    production_max_rows = parse_production_max(excel_file, requested_year=start_year)
    production_forecast_rows = parse_production_forecast(excel_file, requested_year=start_year)

    info["start_year"] = start_year
    return info, sourcing_forecast_rows, production_max_rows, production_forecast_rows, requested_quota_rows


def parse_info(excel_file: Workbook):
    try:
        presentation = excel_file["Présentation"]
        application = excel_file["Reconnaissance double comptage"]

        production_site = presentation[5][2].value
        producer_email = presentation[16][2].value
        try:
            # loop through reconaissance sheet to find the base year defined in it
            year_row_index = 0
            for i, row in enumerate(application.iter_rows()):
                year_row_index = i
                if row[6].value == "Première année de reconnaissance":
                    break
            start_year = int(application[year_row_index + 2][6].value)
        except:
            start_year = 0

        return {
            "production_site": production_site,
            "producer_email": producer_email,
            "start_year": start_year,
        }
    except:
        traceback.print_exc()
        return {"production_site": None, "producer_email": None, "start_year": 0}


def parse_sourcing_forecast(excel_file: Workbook, requested_year: int) -> List[SourcingRow]:
    sourcing_sheet = excel_file["Approvisionnement prévisionnel"]
    sourcing_rows: List[SourcingRow] = []

    current_year = -1

    for line, row in enumerate(sourcing_sheet.iter_rows()):
        current_year = extract_year(row[1].value, current_year)
        if current_year < requested_year:
            continue

        feedstock_name = row[2].value
        origin_country_cell = row[3].value
        supply_country_cell = row[4].value
        transit_country_cell = row[5].value

        # skip row if no year or feedstock is defined
        # TO DELETE : if not feedstock_name or not origin_country_cell or feedstock_name == origin_country_cell:
        if not feedstock_name or feedstock_name == origin_country_cell:
            continue

        feedstock = get_feedstock_from_dc_feedstock(feedstock_name)

        # skip row if no feedstock is recognized and no origin country is defined
        if not feedstock and not origin_country_cell:
            continue

        # this allow to accept row without year but only when feedstock recognized
        if current_year == -1 and feedstock is None:
            continue

        origin_country = extract_country_code(origin_country_cell)
        supply_country = extract_country_code(supply_country_cell)
        transit_country = extract_country_code(transit_country_cell)

        sourcing: SourcingRow = {
            "line": line + 1,
            "year": current_year,
            "feedstock": feedstock,
            "origin_country": origin_country,
            "supply_country": supply_country,
            "transit_country": transit_country,
            "metric_tonnes": 0,
        }

        sourcing["metric_tonnes"] = row[6].value

        sourcing_rows.append(sourcing)

    return sourcing_rows


def intOrZero(value):
    try:
        return int(value)
    except:
        return 0


def parse_production_max(excel_file: Workbook, requested_year) -> List[ProductionMaxRow]:
    production_max_rows: List[ProductionMaxRow] = []
    current_year = -1
    production_sheet = excel_file["Production"]
    for line, row in enumerate(production_sheet.iter_rows()):
        current_year = extract_year(row[1].value, current_year)
        if current_year < requested_year:
            continue

        biofuel_name = None if "SOMME :" == row[2].value else row[2].value
        feedstock_name = row[3].value

        if current_year == -1 or (not feedstock_name and not biofuel_name):
            continue
        feedstock = get_feedstock_from_dc_feedstock(feedstock_name)
        biofuel = get_biofuel_from_dc_biofuel(biofuel_name)

        max_production_capacity = intOrZero(row[4].value)
        production: ProductionMaxRow = {
            "line": line + 1,
            "year": current_year,
            "feedstock": feedstock,
            "biofuel": biofuel,
            "max_production_capacity": max_production_capacity,
        }

        production_max_rows.append(production)

    return production_max_rows


def parse_production_forecast(excel_file: Workbook, requested_year) -> List[ProductionForecastRow]:
    production_forecast_rows: List[ProductionForecastRow] = []

    current_year = -1
    production_sheet = excel_file["Production"]

    for line, row in enumerate(production_sheet.iter_rows()):
        current_year = extract_year(row[6].value, current_year)

        if current_year < requested_year:
            continue
        biofuel_name = None if "SOMME :" == row[7].value else row[7].value
        feedstock_name = row[8].value

        if current_year == -1 or (not feedstock_name and not biofuel_name):
            continue

        feedstock = get_feedstock_from_dc_feedstock(feedstock_name)
        biofuel = get_biofuel_from_dc_biofuel(biofuel_name)
        estimated_production = intOrZero(row[9].value)

        production: ProductionForecastRow = {
            "line": line + 1,
            "year": current_year,
            "feedstock": feedstock,
            "biofuel": biofuel,
            "estimated_production": estimated_production,
        }

        production_forecast_rows.append(production)

    return production_forecast_rows


def parse_requested_quota(excel_file: Workbook) -> List[RequestedQuotaRow]:
    requested_quota_rows: List[RequestedQuotaRow] = []

    current_year = -1
    production_sheet = excel_file["Reconnaissance double comptage"]

    for line, row in enumerate(production_sheet.iter_rows()):
        current_year = extract_year(row[1].value, current_year)

        biofuel_name = None if "SOMME :" == row[2].value else row[2].value
        feedstock_name = row[3].value
        requested_quota = row[4].value

        if current_year == -1 or not requested_quota or (not feedstock_name and not biofuel_name):
            continue

        feedstock = get_feedstock_from_dc_feedstock(feedstock_name)
        biofuel = get_biofuel_from_dc_biofuel(biofuel_name)
        production: RequestedQuotaRow = {
            "line": line + 1,
            "year": current_year,
            "feedstock": feedstock,
            "biofuel": biofuel,
            "requested_quota": requested_quota,
        }

        requested_quota_rows.append(production)

    return requested_quota_rows


def extract_year(year_str: str, current_year: int):
    try:
        match = re.search("([0-9]{4})", str(year_str))
        year = match.group(0)
        return int(year)
    except:
        return current_year


def extract_country_code(country_str: str) -> str | None:
    if country_str:
        return (country_str or "").split("-")[0].strip()
    else:
        return None


def get_feedstock_from_dc_feedstock(feedstock_name: str) -> str | None:
    if not feedstock_name:
        return None
    feedstock_name = feedstock_name.replace("’", "'").strip()
    return dc_feedstock_to_carbure_feedstock.get(feedstock_name, None)


def get_biofuel_from_dc_biofuel(biofuel_name: str) -> str | None:
    if not biofuel_name:
        return None
    biofuel_name = biofuel_name.replace("’", "'").strip()
    return dc_biofuel_to_carbure_biofuel.get(biofuel_name, None)


dc_feedstock_to_carbure_feedstock: dict[str, str | None] = {
    "Algues": "ALGUES",
    "Bagasse": "BAGASSE",
    "Balles (enveloppes)": "BALLES",
    "Betterave": "BETTERAVE",
    "Blé": "BLE",
    "Boues de stations d'épuration": "BOUES_EPURATION",
    "Brai de tallol": "BRAI_TALLOL",
    "Canne à sucre": "CANNE_A_SUCRE",
    "Captage de carbone": None,
    "Colza": "COLZA",
    "Coques": "COQUES",
    "Déchets de bois": "DECHETS_BOIS",
    "Déchets industriels": "DECHETS_INDUSTRIELS",
    "Déchets municipaux en mélange (Hors déchets ménagers triés)": "DECHETS_MUNICIPAUX_MELANGE",
    "Déchets organiques ménagers": "DECHETS_ORGANIQUES_MENAGERS",
    "Distillat d'acide gras de palme": None,
    "Effluents d'huileries de palme et rafles": "EFFLUENTS_HUILERIES_PALME_RAFLE",
    "Effluents d'huileries de palme et rafles (POME)": "EFFLUENTS_HUILERIES_PALME_RAFLE",
    "Egouts Pauvres de 2e Extractions": "EP2",
    "Fumier humide": "FUMIER_HUMIDE",
    "Fumier sec": "FUMIER_SEC",
    "Glycérine brute": "FUMIER_HUMIDE",
    "Graisses de flotation": "GRAISSES_FLOTTATION",
    "Huile alimentaire usagée": "HUILE_ALIMENTAIRE_USAGEE",
    "Huile de palme": "HUILE_PALME",
    "Huiles ou graisses animales (C I)": "HUILES_OU_GRAISSES_ANIMALES_CAT1_CAT2",
    "Huiles ou graisses animales (C II)": "HUILES_OU_GRAISSES_ANIMALES_CAT1_CAT2",
    "Huiles ou graisses animales (C III)": "HUILES_OU_GRAISSES_ANIMALES_CAT3",
    "Lies de vin": "LIES_DE_VIN",
    "Maïs": "MAIS",
    "Marcs de raisin": "MARC_DE_RAISIN",
    "Mat. cellulosiques d'origine non alimentaire": "MAT_CELLULOSIQUE_NON_ALIMENTAIRE",
    "Mat. ligno-cellulosiques (Hors grumes de sciage & de placage)": "MAT_LIGNO_CELLULOSIQUE",
    "Orge": "ORGE",
    "Paille": "PAILLE",
    "Râpes": "RAPES",
    "Seigle": "SEIGLE",
    "Soja": "SOJA",
    "Tallol": "TALLOL",
    "Tournesol": "TOURNESOL",
    "Triticale": "TRITICALE",
    "Distillat d'acide gras de palme": None,
}


dc_biofuel_to_carbure_biofuel: dict[str, str | None] = {
    "Bio Iso-Butène": None,
    "Bio Iso-Octane": None,
    "Bio-essence de synthèse": None,
    "Bio-ETBE": "ETBE",
    "Biogazole de synthèse": "BG",
    "EEHA": "EEHA",
    "EEHU": "EEHU",
    "EEHV": "EEHV",
    "EMAG de POME": "EMAG",
    "EMAG": "EMAG",
    "EMHA": "EMHA",
    "EMHU": "EMHU",
    "EMHV": "EMHV",
    "ETBE": "ETBE",
    "Ethanol d'EP2": "ETH",
    "Ethanol pour ED95": "ED95",
    "Ethanol": "ETH",
    "HVO-C": "HVOC",
    "HVO-E": "HVOE",
    "HVO-G": "HVOG",
    "HC-C": "HCC",
    "HC-E": "HCE",
    "HC-G": "HCG",
    "Méthanol": "MT",
    "MTBE": "MTBE",
    "TAEE": "TAEE",
    "TAME": "TAME",
}
