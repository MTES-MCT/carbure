import time

import openpyxl
from django.http import HttpResponse
from drf_spectacular.utils import OpenApiParameter, extend_schema
from openpyxl.utils import get_column_letter
from rest_framework.decorators import action
from rest_framework.viewsets import GenericViewSet

from tiruert.filters import MacFilter
from tiruert.models import MacFossilFuel


@extend_schema(
    parameters=[
        OpenApiParameter(
            name="entity_id",
            type=int,
            location=OpenApiParameter.QUERY,
            description="Authorised entity ID.",
            required=True,
        ),
    ]
)
class MacFossilFuelExportViewSet(GenericViewSet):
    queryset = MacFossilFuel.objects.all()
    filterset_class = MacFilter

    @action(detail=False, methods=["get"], url_path="export")
    def export_macfossilfuel_to_excel(self, request, *args, **kwargs):
        macs = self.filter_queryset(self.get_queryset())

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "MAC Fossil Fuel Export"

        headers = [
            "SIREN",
            "Opérateur",
            "Carburant",
            "Volume (L)",
            "Année",
        ]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        for row_num, mac in enumerate(macs, 2):
            sheet.cell(row=row_num, column=1, value=mac.operator.registration_id if mac.operator else "")
            sheet.cell(row=row_num, column=2, value=mac.operator.name if mac.operator else "")
            sheet.cell(row=row_num, column=3, value=mac.fuel.nomenclature if mac.fuel else "")
            sheet.cell(row=row_num, column=4, value=mac.volume)
            sheet.cell(row=row_num, column=5, value="2023")

        for col_num in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_num)].width = 20

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"mac_{request.entity}_{time.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        workbook.save(response)
        return response
