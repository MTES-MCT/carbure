import time

import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from rest_framework.decorators import action


class ExcelExportActionMixin:
    @action(
        detail=False,
        methods=["get"],
        url_path="export",
    )
    def export_operations_to_excel(self, request, *args, **kwargs):
        operations = self.filter_queryset(self.get_queryset())

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Operations Export"

        # Header row
        headers = [
            "Statut",
            "Filière",
            "Biocarburant",
            "Catégorie",
            "Date de création",
            "Période de durabilité",
            "Dépôt",
            "Type Opération",
            "Expéditeur",
            "Destinataire",
            "Quantité (L)",
            "Quantité (MJ)",
            "Tonnes CO2 eq. évitées",
        ]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Data rows
        for row_num, operation in enumerate(operations, 2):
            sheet.cell(row=row_num, column=1, value=operation.status)
            sheet.cell(row=row_num, column=2, value=operation.sector)
            sheet.cell(row=row_num, column=3, value=operation.biofuel.code)
            sheet.cell(row=row_num, column=4, value=operation.customs_category)
            sheet.cell(row=row_num, column=5, value=operation.created_at.strftime("%Y-%m-%d"))
            sheet.cell(row=row_num, column=6, value=operation.durability_period)
            sheet.cell(row=row_num, column=7, value=operation._depot if operation._depot else "")
            sheet.cell(row=row_num, column=8, value=operation._type)
            sheet.cell(row=row_num, column=9, value=operation.debited_entity.name if operation.debited_entity else "")
            sheet.cell(row=row_num, column=10, value=operation.credited_entity.name if operation.credited_entity else "")
            sheet.cell(row=row_num, column=11, value=operation._volume)
            sheet.cell(row=row_num, column=12, value=operation.volume_to_quantity(operation._volume, "mj"))
            sheet.cell(row=row_num, column=13, value=sum(detail.avoided_emissions for detail in operation.details.all()))

        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_num)].width = 20

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"tiruert_operations_{request.entity}_{time.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        workbook.save(response)

        return response
