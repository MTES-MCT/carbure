from csv import excel
from imp import get_frozen_object
from typing import List
from openpyxl import Workbook, load_workbook


def extract_country_code(country_str: str) -> str | None:
    if country_str:
        return (country_str or "").split("-")[0].strip()
    else:
        return None


dc_feedstock_to_carbure_feedstock = {
    "Algues": "ALGUES",
    "Bagasse": "BAGASSE",
    "Balles (enveloppes)": "BALLES",
    "Betterave": "BETTERAVE",
    "Blé": "BLE",
    "Boues de stations d’épuration": "BOUES_EPURATION",
    "Brai de tallol": "BRAI_TALLOL",
    "Canne à sucre": "CANNE_A_SUCRE",
    "Captage de carbone": None,
    "Colza": "COLZA",
    "Coques": "COQUES",
    "Déchets de bois": "DECHETS_BOIS",
    "Déchets industriels": "DECHETS_INDUSTRIELS",
    "Déchets municipaux en mélange (Hors déchets ménagers triés)": "DECHETS_MUNICIPAUX_MELANGE",
    "Déchets organiques ménagers": "DECHETS_ORGANIQUES_MENAGERS",
    "Distillat d'acide gras de palme": None,
    "Effluents d’huileries de palme et rafles": "EFFLUENTS_HUILERIES_PALME_RAFLE",
    "Egouts Pauvres de 2e Extractions": "EP2",
    "Fumier humide": "FUMIER_HUMIDE",
    "Fumier sec": "FUMIER_SEC",
    "Glycérine brute": "FUMIER_HUMIDE",
    "Graisses de flotation": "GRAISSES_FLOTTATION",
    "Huile alimentaire usagée": "HUILE_ALIMENTAIRE_USAGEE",
    "Huile de palme": "HUILE_PALME",
    "Huiles ou graisses animales (C I)": "HUILES_OU_GRAISSES_ANIMALES_CAT1_CAT2",
    "Huiles ou graisses animales (C II)": "HUILES_OU_GRAISSES_ANIMALES_CAT1_CAT2",
    "Huiles ou graisses animales (C III)": "HUILES_OU_GRAISSES_ANIMALES_CAT3",
    "Lies de vin": "LIES_DE_VIN",
    "Maïs": "MAIS",
    "Marcs de raisin": "MARC_DE_RAISIN",
    "Mat. cellulosiques d’origine non alimentaire": "MAT_CELLULOSIQUE_NON_ALIMENTAIRE",
    "Mat. ligno-cellulosiques (Hors grumes de sciage & de placage)": "MAT_LIGNO_CELLULOSIQUE",
    "Orge": "ORGE",
    "Paille": "PAILLE",
    "Râpes": "RAPES",
    "Seigle": "SEIGLE",
    "Soja ": "SOJA ",
    "Tallol": "TALLOL",
    "Tournesol ": "TOURNESOL ",
    "Triticale": "TRITICALE",
}

dc_biofuel_to_carbure_biofuel = {
    "Bio Iso-Butène": None,
    "Bio Iso-Octane": None,
    "Bio-essence de synthèse": None,
    "Bio-ETBE": "ETBE",
    "Biogazole de synthèse": "BG",
    "EEHA": None,
    "EEHU": None,
    "EEHV": None,
    "EMAG de POME": "EMAG",
    "EMAG": "EMAG",
    "EMHA": "EMHA",
    "EMHU": "EMHU",
    "EMHV": "EMHV",
    "ETBE": "ETBE",
    "Ethanol d'EP2": "ETH",
    "Ethanol pour ED95": "ED95",
    "Ethanol": "ETH",
    "HVO-C": "HVOC",
    "HVO-E": "HVOE",
    "HVO-G": "HVOG",
    "Méthanol": "MT",
    "MTBE": "MTBE",
    "TAEE": "TAEE",
    "TAME": None,
}

excel_file = load_workbook("/Users/falxa/Desktop/dc.xlsx", data_only=True)


def parse_sourcing(excel_file: Workbook, sheet_name: str) -> List[object]:
    sourcing_sheeet = excel_file[sheet_name]
    sourcing_rows: List[object] = []

    current_year = -1

    for row in sourcing_sheeet.iter_rows():
        try:
            year = int(row[1].value)
            if year != current_year:
                current_year = year
        except Exception:
            pass

        feedstock_name = row[2].value
        origin_country = row[3].value
        supply_country = row[4].value
        transit_country = row[5].value

        # skip row if no year or feedstock is defined
        if current_year == -1 or not feedstock_name or not origin_country:
            continue

        feedstock = dc_feedstock_to_carbure_feedstock.get(feedstock_name, None)
        origin_country_code = extract_country_code(origin_country)
        supply_country_code = extract_country_code(supply_country)
        transit_country_code = extract_country_code(transit_country)

        sourcing_row = {
            "year": current_year,
            "feedstock": feedstock,
            "origin_country": origin_country_code,
            "supply_country": supply_country_code,
            "transit_country": transit_country_code,
        }

        if sheet_name == "Approvisionnement prévisionnel":
            sourcing_row["metric_tonnes"] = row[8].value
        elif sheet_name == "Historique dapprovisionnement":
            sourcing_row["supplier_name"] = row[6].value
            sourcing_row["supplier_certificate"] = row[7].value
            sourcing_row["metric_tonnes"] = row[8].value

        sourcing_rows.append(sourcing_row)

    return sourcing_rows


def parse_production(excel_file: Workbook) -> List[object]:
    production_sheet = excel_file["Production"]
    production_rows: List[object] = []

    current_year = -1

    for row in production_sheet.iter_rows():
        try:
            year = int(row[1].value)
            if year != current_year:
                current_year = year
        except Exception:
            pass

        biofuel_name = row[2].value
        feedstock_name = row[3].value
        max_production_capacity = row[4].value
        estimated_production = row[9].value

        if current_year == -1 or not feedstock_name or not biofuel_name:
            continue

        production_row = {
            "year": current_year,
            "feedstock": dc_feedstock_to_carbure_feedstock.get(feedstock_name, None),
            "biofuel": dc_biofuel_to_carbure_biofuel.get(biofuel_name, None),
            "max_production_capacity": max_production_capacity,
            "estimated_production": estimated_production,
        }

        production_rows.append(production_row)

    quota_sheet = excel_file["Reconnaissance double comptage"]

    current_year = -1

    for row in quota_sheet.iter_rows():
        try:
            year = int(row[1].value)
            if year != current_year:
                current_year = year
        except Exception:
            pass

        biofuel = dc_biofuel_to_carbure_biofuel.get(row[2].value, None)
        feedstock = dc_feedstock_to_carbure_feedstock.get(row[3].value, None)
        requested_quota = row[4].value

        for production in production_rows:
            if (
                current_year == production["year"]
                and biofuel == production["biofuel"]
                and feedstock == production["feedstock"]
            ):
                production["requested_quota"] = requested_quota

    print(production_rows)

    return production_rows


sourcing_forecast = parse_sourcing(excel_file, "Approvisionnement prévisionnel")
sourcing_history = parse_sourcing(excel_file, "Historique dapprovisionnement")
production = parse_production(excel_file)
